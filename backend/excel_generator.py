import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime

def deduplicate_issues_by_row(issues: list[dict]) -> list[dict]:
    """
    Deduplicates issue records by row_index so that each source record
    appears EXACTLY ONCE per sheet, combining multiple issue flags with semicolons.
    """
    grouped = {}
    prio_map = {'Critical': 1, 'High': 2, 'Medium': 3, 'Low': 4}
    
    for issue in issues:
        r_idx = issue['row_index']
        if r_idx not in grouped:
            grouped[r_idx] = {
                'row_index': r_idx,
                'issue_types': [issue['issue_type']],
                'field_names': [issue['field_name']] if issue['field_name'] else [],
                'original_values': [str(issue['original_value'])] if str(issue['original_value']) else [],
                'suggested_corrections': [issue['suggested_correction']],
                'priorities': [issue['priority']],
                'responsible_depts': [issue['responsible_dept']],
                'category': issue['category']
            }
        else:
            if issue['issue_type'] not in grouped[r_idx]['issue_types']:
                grouped[r_idx]['issue_types'].append(issue['issue_type'])
            if issue['field_name'] and issue['field_name'] not in grouped[r_idx]['field_names']:
                grouped[r_idx]['field_names'].append(issue['field_name'])
            if str(issue['original_value']) and str(issue['original_value']) not in grouped[r_idx]['original_values']:
                grouped[r_idx]['original_values'].append(str(issue['original_value']))
            if issue['suggested_correction'] not in grouped[r_idx]['suggested_corrections']:
                grouped[r_idx]['suggested_corrections'].append(issue['suggested_correction'])
            if issue['priority'] not in grouped[r_idx]['priorities']:
                grouped[r_idx]['priorities'].append(issue['priority'])
            if issue['responsible_dept'] not in grouped[r_idx]['responsible_depts']:
                grouped[r_idx]['responsible_depts'].append(issue['responsible_dept'])

    deduped = []
    for r_idx in sorted(grouped.keys()):
        item = grouped[r_idx]
        sorted_prios = sorted(item['priorities'], key=lambda p: prio_map.get(p, 5))
        top_prio = sorted_prios[0] if sorted_prios else 'Medium'
        
        deduped.append({
            'row_index': r_idx,
            'issue_type': '; '.join(item['issue_types']),
            'field_name': '; '.join(item['field_names']),
            'original_value': '; '.join(item['original_values']),
            'suggested_correction': '; '.join(item['suggested_corrections']),
            'priority': top_prio,
            'responsible_dept': '; '.join(item['responsible_depts']),
            'category': item['category']
        })
    return deduped

def deduplicate_category_issues_distinctly(issues: list[dict], df: pd.DataFrame, id_col: str) -> list[dict]:
    """
    Filters out redundant duplicate rows and deduplicates category warnings by (primary_key, issue_type, field_name)
    so that duplicate rows in the input file do not repeat identical warnings in non-duplicate sheets.
    """
    seen_keys = set()
    filtered = []
    
    for issue in issues:
        # Skip exact duplicate row flags in category sheets (handled in Duplicate Records sheet)
        if issue['issue_type'] == 'Exact Duplicate Row':
            continue
            
        orig_row_idx = issue['row_index'] - 1
        pk_val = ""
        if id_col and orig_row_idx < len(df) and id_col in df.columns:
            pk_val = str(df.iloc[orig_row_idx][id_col]).strip()
            
        dedup_key = (pk_val or orig_row_idx, issue['category'], issue['issue_type'], issue['field_name'], str(issue['original_value']).strip())
        
        if dedup_key not in seen_keys:
            seen_keys.add(dedup_key)
            filtered.append(issue)
            
    return deduplicate_issues_by_row(filtered)

def generate_excel_workbook(
    df: pd.DataFrame,
    analysis_results: dict,
    filename: str,
    branch: str,
    master_type: str,
    checksum: str,
    output_filepath: str
):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E79")
    subtitle_font = Font(name=font_family, size=11, italic=True, color="595959")
    section_font = Font(name=font_family, size=12, bold=True, color="1F4E79")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=10, bold=True)
    regular_font = Font(name=font_family, size=10)
    link_font = Font(name=font_family, size=10, underline="single", color="0563C1")
    
    navy_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    teal_header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    slate_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(name=font_family, size=10, color="9C0006", bold=True)
    
    amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    amber_font = Font(name=font_family, size=10, color="9C6500", bold=True)
    
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    yellow_font = Font(name=font_family, size=10, color="B25900")
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(name=font_family, size=10, color="006100")
    
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    grey_font = Font(name=font_family, size=10, color="595959")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    id_col = analysis_results.get('id_column', '')

    # -------------------------------------------------------------
    # SHEET 1: EXECUTIVE SUMMARY
    # -------------------------------------------------------------
    ws_exec = wb.create_sheet(title="Executive Summary")
    ws_exec.views.sheetView[0].showGridLines = True
    
    ws_exec.cell(row=1, column=1, value="Kranium HIS Master Sanitization — Executive Summary").font = title_font
    ws_exec.cell(row=2, column=1, value="Prashanth Hospital Master File Quality Audit & Sanitization Report").font = subtitle_font
    
    ws_exec.cell(row=4, column=1, value="File Analysis Metadata").font = section_font
    meta_items = [
        ("Source Filename:", filename),
        ("Analysis Date & Time:", datetime.now().strftime("%d-%b-%Y %H:%M:%S")),
        ("Hospital Branch:", branch),
        ("Master Type:", master_type),
        ("File SHA-256 Checksum:", checksum),
        ("Integrity Notice:", "Original uploaded source file was NOT modified.")
    ]
    for r_idx, (k, v) in enumerate(meta_items, start=5):
        cell_k = ws_exec.cell(row=r_idx, column=1, value=k)
        cell_k.font = bold_font
        cell_v = ws_exec.cell(row=r_idx, column=2, value=v)
        cell_v.font = regular_font
        if k.startswith("Integrity"):
            cell_v.font = Font(name=font_family, size=10, bold=True, color="008000")
            
    ws_exec.cell(row=12, column=1, value="Executive Metrics Breakdown").font = section_font
    metrics_headers = ["Metric Category", "Count", "Description & Interpretation"]
    for c_idx, h in enumerate(metrics_headers, start=1):
        cell = ws_exec.cell(row=13, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx==2 else "left")
        
    metrics_rows = [
        ("Source Records", analysis_results['total_records'], "Total raw rows uploaded from source file"),
        ("Unique Record IDs", analysis_results['unique_records'], "Distinct primary key identifiers present"),
        ("Affected Records", analysis_results['affected_records'], "Unique source records containing 1 or more data issues"),
        ("Clean Records", analysis_results['clean_records'], "Records completely free of identified data errors"),
        ("Total Issue Flags", analysis_results['total_issues'], "Total individual error/warning flags across all records"),
        ("Critical Priority Issues", analysis_results['critical_issues'], "Severe issues requiring immediate correction before HIS load"),
        ("High Priority Issues", analysis_results['high_priority_issues'], "High risk issues requiring department signoff"),
        ("Duplicate Record IDs", analysis_results['duplicate_ids'], "Records sharing identical key identifiers"),
        ("Repeated Rows", analysis_results['repeated_rows'], "100% identical redundant rows"),
        ("Missing Mandatory Fields", analysis_results['missing_fields'], "Empty cells in mandatory primary/attribute columns"),
        ("Invalid Field Values", analysis_results['invalid_values'], "Values failing data-type or controlled list validation"),
        ("Casing & Format Flags", analysis_results.get('casing_issues', 0), "Name casing, digits in names, and login ID formatting errors")
    ]
    
    for r_idx, (cat, val, desc) in enumerate(metrics_rows, start=14):
        c1 = ws_exec.cell(row=r_idx, column=1, value=cat)
        c2 = ws_exec.cell(row=r_idx, column=2, value=val)
        c3 = ws_exec.cell(row=r_idx, column=3, value=desc)
        
        c1.font = bold_font
        c2.font = bold_font
        c3.font = regular_font
        
        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border
        c2.alignment = Alignment(horizontal="center")
        
        if cat in ["Affected Records", "Total Issue Flags", "Critical Priority Issues"]:
            c1.fill = amber_fill if cat != "Critical Priority Issues" else red_fill
            c2.fill = amber_fill if cat != "Critical Priority Issues" else red_fill

    all_issues_list = analysis_results['all_issues']

    category_subsets = [
        ("Duplicate Records", deduplicate_issues_by_row([i for i in all_issues_list if i['category'] == 'Duplicate Records']), "duplicate key and repeated row findings"),
        ("Missing Fields", deduplicate_category_issues_distinctly([i for i in all_issues_list if i['category'] == 'Missing Fields'], df, id_col), "missing mandatory values sheet"),
        ("Invalid Values", deduplicate_category_issues_distinctly([i for i in all_issues_list if i['category'] == 'Invalid Values'], df, id_col), "invalid status and formatting errors"),
        ("Spelling and Standardisation", deduplicate_category_issues_distinctly([i for i in all_issues_list if i['category'] == 'Spelling and Standardisation'], df, id_col), "whitespace, typos, and spelling standardisation findings"),
        ("Casing and Formatting", deduplicate_category_issues_distinctly([i for i in all_issues_list if i['category'] == 'Casing and Formatting'], df, id_col), "name casing, digits in names, and login ID formatting findings"),
        ("Department-Specific Review", deduplicate_category_issues_distinctly([i for i in all_issues_list if i['category'] == 'Department-Specific Review'], df, id_col), "clinical / drug / tariff safety warnings")
    ]

    ws_exec.cell(row=28, column=1, value="Quick Sheet Navigation Links").font = section_font
    nav_links = [
        ("View Original Master Data", "'Original Master Data'!A1", "Go to 100% complete raw uploaded source dataset", True),
        ("View Affected Records", "'Affected Records'!A1", "Go to list of unique records requiring sanitization", True),
        ("View Critical Issues", "'All Issues'!A1", "Go to master issues log containing all flags", True),
    ]
    
    for cat_name, issues_sub, purp in category_subsets:
        if len(issues_sub) > 0:
            nav_links.append((f"View {cat_name}", f"'{cat_name}'!A1", f"Go to {purp} ({len(issues_sub)} affected records)", True))
        else:
            nav_links.append((f"View {cat_name}", None, f"100% Compliant — 0 issues detected ({cat_name} sheet omitted)", False))

    for c_idx, h in enumerate(["Target Sheet", "Hyperlink Action", "Sheet Purpose"], start=1):
        c = ws_exec.cell(row=29, column=c_idx, value=h)
        c.font = header_font
        c.fill = teal_header_fill
        
    for r_idx, (label, target, purp, has_link) in enumerate(nav_links, start=30):
        c1 = ws_exec.cell(row=r_idx, column=1, value=label)
        if has_link and target:
            c2 = ws_exec.cell(row=r_idx, column=2, value=f"=HYPERLINK(\"#{target}\", \"Open Sheet ->\")")
            c2.font = link_font
        else:
            c2 = ws_exec.cell(row=r_idx, column=2, value="Compliant (0 Issues)")
            c2.font = grey_font
            
        c3 = ws_exec.cell(row=r_idx, column=3, value=purp)
        
        c1.font = bold_font
        c3.font = regular_font
        
        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border

    ws_exec.cell(row=40, column=1, value="Correction Priorities & Workstream Roadmap").font = section_font
    prio_headers = ["Order", "Workstream", "Evidence", "Priority", "Required Action", "Responsible Department"]
    for c_idx, h in enumerate(prio_headers, start=1):
        c = ws_exec.cell(row=41, column=c_idx, value=h)
        c.font = header_font
        c.fill = navy_header_fill
        
    for r_idx, item in enumerate(analysis_results['priorities_list'], start=42):
        c1 = ws_exec.cell(row=r_idx, column=1, value=item['order'])
        c2 = ws_exec.cell(row=r_idx, column=2, value=item['workstream'])
        c3 = ws_exec.cell(row=r_idx, column=3, value=item['evidence'])
        c4 = ws_exec.cell(row=r_idx, column=4, value=item['priority'])
        c5 = ws_exec.cell(row=r_idx, column=5, value=item['required_action'])
        c6 = ws_exec.cell(row=r_idx, column=6, value=item['responsible_dept'])
        
        for c in [c1, c2, c3, c4, c5, c6]:
            c.font = regular_font
            c.border = thin_border
        c1.alignment = Alignment(horizontal="center")
        c4.font = bold_font
        if item['priority'] == 'Critical':
            c4.fill = red_fill
            c4.font = red_font
        elif item['priority'] == 'High':
            c4.fill = amber_fill
            c4.font = amber_font

    # -------------------------------------------------------------
    # SHEET 2: ORIGINAL MASTER DATA (100% Raw Uploaded Source Data)
    # -------------------------------------------------------------
    ws_orig = wb.create_sheet(title="Original Master Data")
    ws_orig.views.sheetView[0].showGridLines = True
    ws_orig.cell(row=1, column=1, value='=HYPERLINK("#\'Executive Summary\'!A1", "<- Back to Executive Summary")').font = link_font
    
    orig_cols = list(df.columns)
    orig_headers = orig_cols + ["Data Audit Status"]
    
    for c_idx, h in enumerate(orig_headers, start=1):
        c = ws_orig.cell(row=3, column=c_idx, value=h)
        c.font = header_font
        c.fill = navy_header_fill if c_idx <= len(orig_cols) else teal_header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_orig.row_dimensions[3].height = 28
    
    for r_idx in range(len(df)):
        row_data = df.iloc[r_idx]
        for c_idx, c_name in enumerate(orig_cols, start=1):
            c = ws_orig.cell(row=r_idx+4, column=c_idx, value=str(row_data.get(c_name, "")))
            c.font = regular_font
            c.border = thin_border
        c_st = ws_orig.cell(row=r_idx+4, column=len(orig_cols)+1, value="Original Source Data")
        c_st.font = regular_font
        c_st.fill = slate_fill
        c_st.border = thin_border
        
    ws_orig.freeze_panes = 'A4'
    ws_orig.auto_filter.ref = f"A3:{get_column_letter(len(orig_headers))}{len(df)+3}"

    # -------------------------------------------------------------
    # CORRECTION SHEETS (DYNAMIC CATEGORY SHEETS)
    # -------------------------------------------------------------
    governance_headers = [
        "Issue Type", "Field Name", "Original Value", "Suggested Correction/Action",
        "Corrected Value", "Priority", "Responsible Department", "Correction Status",
        "Corrected By", "Correction Reason", "Remarks"
    ]
    
    all_headers = orig_cols + governance_headers
    dv_status = DataValidation(type="list", formula1='"Pending,Corrected,No Change Required,Needs Clarification"', allow_blank=True)
    
    sheet_definitions = [
        ("All Issues", deduplicate_category_issues_distinctly(all_issues_list, df, id_col)),
        ("Affected Records", deduplicate_category_issues_distinctly(all_issues_list, df, id_col)),
    ]
    for cat_name, issues_sub, _ in category_subsets:
        if len(issues_sub) > 0:
            sheet_definitions.append((cat_name, issues_sub))

    for sheet_name, issues_subset in sheet_definitions:
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        ws.add_data_validation(dv_status)
        
        link_cell = ws.cell(row=1, column=1, value='=HYPERLINK("#\'Executive Summary\'!A1", "<- Back to Executive Summary")')
        link_cell.font = link_font
        
        for col_idx, h in enumerate(all_headers, start=1):
            c = ws.cell(row=3, column=col_idx, value=h)
            c.font = header_font
            c.fill = navy_header_fill if col_idx <= len(orig_cols) else teal_header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        ws.row_dimensions[3].height = 28
        row_counter = 4
        
        # If issues_subset is empty for this category, display 1 clean green compliance summary row
        if not issues_subset:
            for col_idx in range(1, len(orig_cols) + 1):
                c = ws.cell(row=4, column=col_idx, value="N/A")
                c.font = regular_font
                c.border = thin_border
                
            no_issue_vals = [
                "Clean / Compliant",
                "ALL_FIELDS",
                "Compliant",
                "No action required — 100% compliant in this check category.",
                "",
                "Low",
                master_type,
                "No Change Required",
                "System Validator",
                "Passes Validation",
                "Zero issues detected in this category"
            ]
            
            for g_offset, g_val in enumerate(no_issue_vals):
                c_col = len(orig_cols) + 1 + g_offset
                c = ws.cell(row=4, column=c_col, value=g_val)
                c.font = regular_font
                c.border = thin_border
                if g_offset == 7: # Correction Status
                    c.fill = green_fill
                    c.font = green_font
                    dv_status.add(c)
            row_counter = 5
        else:
            # Build Data Rows for Issues Found in this specific category (Deduplicated per source row)
            for issue in issues_subset:
                orig_row_idx = issue['row_index'] - 1
                row_data = df.iloc[orig_row_idx] if orig_row_idx < len(df) else {}
                
                # 1. Fill Original Source Columns for this affected row
                for col_idx, c_name in enumerate(orig_cols, start=1):
                    val = row_data.get(c_name, "") if isinstance(row_data, pd.Series) else ""
                    c = ws.cell(row=row_counter, column=col_idx, value=str(val))
                    c.font = regular_font
                    c.border = thin_border
                    
                # 2. Fill Governance Added Columns
                gov_vals = [
                    issue['issue_type'],
                    issue['field_name'],
                    issue['original_value'],
                    issue['suggested_correction'],
                    "", # Corrected Value
                    issue['priority'],
                    issue['responsible_dept'],
                    "Pending", # Default Correction Status
                    "", # Corrected By
                    "", # Correction Reason
                    ""  # Remarks
                ]
                
                for g_offset, g_val in enumerate(gov_vals):
                    c_col = len(orig_cols) + 1 + g_offset
                    c = ws.cell(row=row_counter, column=c_col, value=g_val)
                    c.font = regular_font
                    c.border = thin_border
                    
                    if g_offset == 5:
                        if g_val == 'Critical':
                            c.fill = red_fill
                            c.font = red_font
                        elif g_val == 'High':
                            c.fill = amber_fill
                            c.font = amber_font
                        elif g_val == 'Medium':
                            c.fill = yellow_fill
                            c.font = yellow_font
                            
                    if g_offset == 7:
                        c.fill = grey_fill
                        c.font = grey_font
                        dv_status.add(c)
                        
                row_counter += 1
            
        ws.freeze_panes = 'A4'
        ws.auto_filter.ref = f"A3:{get_column_letter(len(all_headers))}{row_counter-1}"

    # -------------------------------------------------------------
    # SHEET 11: CORRECTION PRIORITIES
    # -------------------------------------------------------------
    ws_cp = wb.create_sheet(title="Correction Priorities")
    ws_cp.views.sheetView[0].showGridLines = True
    ws_cp.cell(row=1, column=1, value='=HYPERLINK("#\'Executive Summary\'!A1", "<- Back to Executive Summary")').font = link_font
    
    cp_headers = ["Priority Rank", "Sanitization Workstream", "Found Evidence", "Assigned Priority", "Action Required", "Responsible Team"]
    for c_idx, h in enumerate(cp_headers, start=1):
        c = ws_cp.cell(row=3, column=c_idx, value=h)
        c.font = header_font
        c.fill = navy_header_fill
        
    for r_idx, item in enumerate(analysis_results['priorities_list'], start=4):
        ws_cp.cell(row=r_idx, column=1, value=item['order']).alignment = Alignment(horizontal="center")
        ws_cp.cell(row=r_idx, column=2, value=item['workstream'])
        ws_cp.cell(row=r_idx, column=3, value=item['evidence'])
        c_p = ws_cp.cell(row=r_idx, column=4, value=item['priority'])
        ws_cp.cell(row=r_idx, column=5, value=item['required_action'])
        ws_cp.cell(row=r_idx, column=6, value=item['responsible_dept'])
        
        for c_i in range(1, 7):
            c_cell = ws_cp.cell(row=r_idx, column=c_i)
            c_cell.border = thin_border
            c_cell.font = regular_font
            
        if item['priority'] == 'Critical':
            c_p.fill = red_fill
            c_p.font = red_font
        elif item['priority'] == 'High':
            c_p.fill = amber_fill
            c_p.font = amber_font
            
    ws_cp.freeze_panes = 'A4'

    # -------------------------------------------------------------
    # SHEET 12: SANITIZATION RULES
    # -------------------------------------------------------------
    ws_rules = wb.create_sheet(title="Sanitization Rules")
    ws_rules.views.sheetView[0].showGridLines = True
    ws_rules.cell(row=1, column=1, value='=HYPERLINK("#\'Executive Summary\'!A1", "<- Back to Executive Summary")').font = link_font
    
    rule_headers = ["Rule ID", "Category", "Validation Check Name", "Target Columns", "Threshold / Logic", "Safety Protocol"]
    for c_idx, h in enumerate(rule_headers, start=1):
        c = ws_rules.cell(row=3, column=c_idx, value=h)
        c.font = header_font
        c.fill = teal_header_fill
        
    rule_rows = [
        ("RULE-01", "Universal", "Primary Key Uniqueness", "Primary Code / ID", "100% Unique Required", "Must not automerge without ID verification"),
        ("RULE-02", "Universal", "Exact Duplicate Row Check", "All Record Columns", "Identical Tuple Check", "Flag identical redundant rows for deletion"),
        ("RULE-03", "Universal", "Whitespace Cleaning", "All String Fields", "Trim leading/trailing/multiple spaces", "Auto-trim formatting whitespace"),
        ("RULE-04", "Universal", "Casing & Title Standardisation", "Name & Description Fields", "Standardize ALL CAPS / lowercase to Title Case", "Flag casing inconsistencies for rectification"),
        ("RULE-05", "Universal", "Login ID Format Audit", "Login ID, User Name", "Lowercase alphanumeric without spaces/symbols", "Flag non-standard login identifiers"),
        ("RULE-06", "Pharmacy", "Generic Salt Mapping", "Drug Name, Generic Code", "Must link to valid generic salt", "Department Validation Required"),
        ("RULE-07", "Pharmacy", "Drug Schedule Validation", "Schedule Type", "OTC, Schedule H, H1, X, Narcotic", "Department Validation Required"),
        ("RULE-08", "Doctors", "Medical Registration Audit", "Registration No, Council", "Valid Council No & Name match", "Department Validation Required"),
        ("RULE-09", "Laboratory", "Specimen & Reference Range", "Specimen, Unit, TAT", "NABL Accredited specimen mapping", "Department Validation Required"),
        ("RULE-10", "Tariffs", "Payer Package Mapping", "Service Code, Tariff Price", "Non-zero price & active billing code", "Department Validation Required")
    ]
    
    for r_idx, r_data in enumerate(rule_rows, start=4):
        for c_idx, val in enumerate(r_data, start=1):
            c = ws_rules.cell(row=r_idx, column=c_idx, value=val)
            c.font = regular_font
            c.border = thin_border
            if c_idx == 1:
                c.alignment = Alignment(horizontal="center")
                c.font = bold_font
                
    ws_rules.freeze_panes = 'A4'

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if not val_str.startswith("=HYPERLINK") and len(val_str) < 80:
                    max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_filepath)
