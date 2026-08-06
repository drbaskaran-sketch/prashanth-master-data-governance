import requests
import io
import openpyxl

print("==========================================================================")
print("🏥 RUNNING COMPLETE END-TO-END SYSTEM INTEGRATION TEST")
print("==========================================================================")

BASE_URL = "http://localhost:5050"
FRONTEND_URL = "http://localhost:3001"

# 1. Test Backend Health
print("\n1. Testing Backend API Health...")
r_health = requests.get(f"{BASE_URL}/api/health")
assert r_health.status_code == 200, f"Health check failed: {r_health.text}"
print("✅ Health Check Passed:", r_health.json())

# 2. Test Frontend Server Assets on Port 3001...
print("\n2. Testing Frontend Server Assets on Port 3001...")
r_html = requests.get(FRONTEND_URL)
assert r_html.status_code == 200, f"Frontend check failed: {r_html.status_code}"
html_text = r_html.text

assert "Prashanth Hospital" in html_text or "Master File Analyzer" in html_text
assert "Master Registry Schemas" not in html_text
assert "Multi-Role Approval Queues" not in html_text
print("✅ Frontend HTML Asset Verification Passed!")

# 3. Test End-to-End User / General Master Upload with Casing & Login ID Rules
print("\n3. Testing User / Employee Master Upload with Casing & Login ID Validation...")
user_master_csv = """user_id,employee_name,login_id,status
EMP-001,JOHN DOE,john.doe,ACTIVE
EMP-002,jane smith,JaneSmith123,ACTIVE
EMP-003,Robert99,robert smith,ACTIVE
EMP-004,  Dr. A. R. Baskaran  ,dr.baskaran,INVALID_STATUS
"""

files = {'file': ('General_User_Master_Sample.csv', user_master_csv, 'text/csv')}
data = {'branch': 'Velachery Unit', 'department': 'General/Other Masters', 'master_type': 'Employee Master'}

r_analyze = requests.post(f"{BASE_URL}/api/analyze", files=files, data=data)
assert r_analyze.status_code == 200, f"Analysis failed: {r_analyze.text}"

result = r_analyze.json()
assert result['success'] == True
summary = result['summary']

print("   Summary Results:")
print(f"   • Total Records:          {summary['total_records']}")
print(f"   • Unique Records:         {summary['unique_records']}")
print(f"   • Affected Records:       {summary['affected_records']}")
print(f"   • Clean Records:          {summary['clean_records']}")
print(f"   • Total Issues:           {summary['total_issues']}")
print(f"   • Casing & Format Flags:  {summary.get('casing_issues', 0)}")

assert summary['total_records'] == 4
assert summary['affected_records'] == 4
assert summary.get('casing_issues', 0) >= 4
print("✅ Casing & Login ID Format Metrics Verification Passed!")

# 4. Test Excel Correction Workbook Download & Structure
print("\n4. Testing Excel Correction Workbook Download & Structure...")
excel_url = f"{BASE_URL}{result['excel_download_url']}"
r_excel = requests.get(excel_url)
assert r_excel.status_code == 200, f"Excel download failed: {r_excel.status_code}"

wb = openpyxl.load_workbook(io.BytesIO(r_excel.content))
expected_sheets = [
    'Executive Summary', 'Original Master Data', 'All Issues', 'Affected Records',
    'Invalid Values', 'Spelling and Standardisation', 'Casing and Formatting',
    'Correction Priorities', 'Sanitization Rules'
]

print("   Workbook Sheet Names:", wb.sheetnames)
assert wb.sheetnames == expected_sheets, f"Sheets mismatch: {wb.sheetnames}"

# Check Casing and Formatting sheet
ws_casing = wb['Casing and Formatting']
assert ws_casing.cell(row=4, column=1).value == "EMP-001"
# Orig cols = 4 (user_id, employee_name, login_id, status), so governance starts at col 5
issue_type_val = ws_casing.cell(row=4, column=5).value
suggested_val = ws_casing.cell(row=4, column=8).value
print(f"   Row 4 -> Issue Type (Col 5): '{issue_type_val}', Suggested (Col 8): '{suggested_val}'")
assert "Numeric Digits" in str(issue_type_val) or "ALL CAPS" in str(issue_type_val) or "Emp" in str(suggested_val) or "John Doe" in str(suggested_val)
print("✅ Casing and Formatting Sheet Row Verification Passed!")

print("✅ Excel Workbook Structure & Link Verification Passed!")

# 5. Test CSV Issues Download
print("\n5. Testing CSV Issues Export Download...")
csv_url = f"{BASE_URL}{result['csv_download_url']}"
r_csv = requests.get(csv_url)
assert r_csv.status_code == 200, f"CSV download failed: {r_csv.status_code}"
assert "Casing" in r_csv.text or "Login" in r_csv.text or "ALL CAPS" in r_csv.text or "Numeric Digits" in r_csv.text
print("✅ CSV Export Verification Passed!")

print("\n==========================================================================")
print("🎉 ALL END-TO-END SYSTEM INTEGRATION TESTS PASSED SUCCESSFULLY!")
print("==========================================================================")
